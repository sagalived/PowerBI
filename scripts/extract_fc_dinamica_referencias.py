import json
import os
import sys
from datetime import datetime

import openpyxl

DEFAULT_EXCEL_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "FC_Maio_Modelo.xlsx")
DEFAULT_SHEET_NAME = "FC_DINÂMICA"
DEFAULT_OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "web",
    "data",
    "referencias_modelo.json",
)


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_sheet_rows(excel_path: str, sheet_name: str):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Aba '{sheet_name}' não encontrada. Abas: {workbook.sheetnames}")

    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column

    rows = []
    for r in range(1, max_row + 1):
        values = [_text(sheet.cell(r, c).value) for c in range(1, max_col + 1)]
        if any(v for v in values):
            rows.append({"row": r, "values": values})

    return {
        "sheet": sheet_name,
        "maxRow": max_row,
        "maxCol": max_col,
        "rows": rows,
    }


def main(argv: list[str]) -> int:
    excel_path = argv[1] if len(argv) > 1 else DEFAULT_EXCEL_PATH
    sheet_name = argv[2] if len(argv) > 2 else DEFAULT_SHEET_NAME
    output_path = argv[3] if len(argv) > 3 else DEFAULT_OUTPUT_PATH

    if not os.path.exists(excel_path):
        raise SystemExit(
            "Arquivo Excel não encontrado. Informe o caminho como 1º argumento. "
            f"Padrão tentado: {excel_path}"
        )

    payload = {
        "sourceFile": os.path.basename(excel_path),
        "sourcePath": excel_path,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        **extract_sheet_rows(excel_path, sheet_name),
    }

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"OK: {output_path}")
    print(f"Rows exportadas: {len(payload['rows'])} (maxRow={payload['maxRow']}, maxCol={payload['maxCol']})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
