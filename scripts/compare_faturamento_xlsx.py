import argparse
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import openpyxl


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = _text(value)
    if not s:
        return 0.0
    # remove moeda e separadores pt-BR
    s = s.replace("R$", "").strip()
    s = s.replace(".", "").replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_key(s: str) -> str:
    s = _text(s).lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _iso_date(s: Any) -> str:
    if isinstance(s, datetime):
        return s.date().isoformat()
    t = _text(s)
    if not t:
        return ""
    # se já está ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}$", t):
        return t
    # tenta dd/mm/aaaa
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", t)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo}-{d}"
    return ""


def _detect_period_from_xlsx(path: str, sheet_name: str | None = None) -> tuple[str, str] | None:
    """Tenta detectar o período configurado no relatório (string 'Período').

    Retorna (inicioISO, fimISO) ou None.
    Exemplo esperado: '01/01/2026 a 30/04/2026'.
    """

    wb = openpyxl.load_workbook(path, data_only=True)
    sh = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb[wb.sheetnames[0]]

    max_row = min(sh.max_row or 1, 80)
    max_col = min(sh.max_column or 1, 60)

    period_re = re.compile(r"(\d{2}/\d{2}/\d{4})\s*a\s*(\d{2}/\d{2}/\d{4})")

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = sh.cell(r, c).value
            if not (isinstance(v, str) and "Período" in v):
                continue

            # normalmente o valor está algumas colunas à direita (ex: col+4)
            for c2 in range(c + 1, min(c + 25, max_col) + 1):
                v2 = sh.cell(r, c2).value
                if not isinstance(v2, str):
                    continue
                m = period_re.search(v2)
                if not m:
                    continue
                ini, fim = m.groups()
                ini_iso = _iso_date(ini)
                fim_iso = _iso_date(fim)
                if ini_iso and fim_iso:
                    return (ini_iso, fim_iso)

    return None


@dataclass
class ExtractedRow:
    key: str
    total: float
    titulo: str = ""
    documento: str = ""
    data: str = ""


def _parse_date_any(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = _text(v)
    if not t:
        return None
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", t)
    if m:
        d, mo, y = m.groups()
        try:
            return date(int(y), int(mo), int(d))
        except ValueError:
            return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", t)
    if m:
        y, mo, d = m.groups()
        try:
            return date(int(y), int(mo), int(d))
        except ValueError:
            return None
    return None


def find_header_grupo_total(sheet) -> tuple[int, int, int] | None:
    """Retorna (headerRow, grupoCol, totalCol) (1-based) ou None."""
    max_row = min(sheet.max_row or 1, 200)
    max_col = min(sheet.max_column or 1, 80)

    def norm_cell(v: Any) -> str:
        return _norm_key(_text(v))

    for r in range(1, max_row + 1):
        values = [norm_cell(sheet.cell(r, c).value) for c in range(1, max_col + 1)]
        if not any(values):
            continue

        grupo_cols = [i + 1 for i, v in enumerate(values) if v in {"grupo", "grupos"}]
        total_cols = [i + 1 for i, v in enumerate(values) if v in {"total", "valor", "valor total"}]
        if not grupo_cols or not total_cols:
            continue

        # escolhe o total mais próximo do primeiro grupo
        gcol = grupo_cols[0]
        tcol = min(total_cols, key=lambda x: abs(x - gcol))
        return (r, gcol, tcol)

    return None


def find_header_faturamento(sheet) -> dict[str, int] | None:
    """Retorna um mapa de colunas (1-based) para o layout 'Faturamento' do Sienge.

    Espera cabeçalhos como:
      - Centro de custo
      - Valor bruto
      - Título (opcional)
      - Documento (opcional)
      - Data emissão (opcional)
    """

    max_row = min(sheet.max_row or 1, 200)
    max_col = min(sheet.max_column or 1, 120)

    def norm_cell(v: Any) -> str:
        return _norm_key(_text(v))

    for r in range(1, max_row + 1):
        values = [norm_cell(sheet.cell(r, c).value) for c in range(1, max_col + 1)]
        if not any(values):
            continue

        # coluna obrigatória: centro de custo + valor bruto
        try:
            cc_col = values.index("centro de custo") + 1
        except ValueError:
            continue

        # "valor bruto" pode variar um pouco
        vb_candidates = {"valor bruto", "valor", "valor (bruto)", "valor bruto (r$)"}
        vb_col = None
        for i, v in enumerate(values, start=1):
            if v in vb_candidates:
                vb_col = i
                break
        if not vb_col:
            continue

        cols: dict[str, int] = {"header_row": r, "centro_custo": cc_col, "valor_bruto": vb_col}

        for key, header in [
            ("cliente", "cliente"),
            ("documento", "documento"),
            ("titulo", "título"),
            ("titulo", "titulo"),
            ("data_emissao", "data emissão"),
            ("data_emissao", "data emissao"),
        ]:
            if header in values:
                cols[key] = values.index(header) + 1

        return cols

    return None


def extract_rows_from_xlsx(
    path: str,
    sheet_name: str | None = None,
    inicio: str | None = None,
    fim: str | None = None,
) -> tuple[str, str, list[ExtractedRow]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for name in sheets:
        if name not in wb.sheetnames:
            continue
        sh = wb[name]

        # 1) tenta layout novo (Faturamento por Centro de Custo)
        cols = find_header_faturamento(sh)
        if cols:
            header_row = cols["header_row"]
            cc_col = cols["centro_custo"]
            vb_col = cols["valor_bruto"]
            cliente_col = cols.get("cliente")
            doc_col = cols.get("documento")
            titulo_col = cols.get("titulo")
            dt_col = cols.get("data_emissao")

            inicio_d = _parse_date_any(inicio) if inicio else None
            fim_d = _parse_date_any(fim) if fim else None

            out: list[ExtractedRow] = []
            empty_streak = 0
            for r in range(header_row + 1, (sh.max_row or header_row) + 1):
                centro = _text(sh.cell(r, cc_col).value)
                valor = _num(sh.cell(r, vb_col).value)

                if not centro and valor == 0:
                    empty_streak += 1
                    if empty_streak >= 50:
                        break
                    continue
                empty_streak = 0

                # pula subtotais
                if _norm_key(centro).startswith("total"):
                    continue

                dt_str = ""
                if dt_col:
                    dt = _parse_date_any(sh.cell(r, dt_col).value)
                    if dt:
                        dt_str = dt.isoformat()
                        if inicio_d and dt < inicio_d:
                            continue
                        if fim_d and dt > fim_d:
                            continue

                titulo = _text(sh.cell(r, titulo_col).value) if titulo_col else ""
                documento = _text(sh.cell(r, doc_col).value) if doc_col else ""
                cliente = _text(sh.cell(r, cliente_col).value) if cliente_col else ""

                # chave default: centro de custo (texto) — mantém o id quando vem no relatório
                key = centro
                if cliente and key and "|" not in key:
                    # evita conflitar centros com mesmo nome em clientes diferentes (raro)
                    pass

                out.append(ExtractedRow(key=key, total=valor, titulo=titulo, documento=documento, data=dt_str))

            if out:
                return (name, "faturamento", out)

        # 2) layout antigo (Grupo/Total)
        header = find_header_grupo_total(sh)
        if not header:
            continue

        header_row, grupo_col, total_col = header

        out2: list[ExtractedRow] = []
        empty_streak = 0
        for r in range(header_row + 1, (sh.max_row or header_row) + 1):
            grupo = _text(sh.cell(r, grupo_col).value)
            total = _num(sh.cell(r, total_col).value)

            if not grupo and total == 0:
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0

            if grupo:
                out2.append(ExtractedRow(key=grupo, total=total))

        if out2:
            return (name, "grupo_total", out2)

    raise SystemExit(
        "Não consegui identificar o layout do XLSX. "
        "Para este comparador eu espero ou (a) colunas 'Centro de custo' + 'Valor bruto' "
        "ou (b) colunas 'Grupo' + 'Total'."
    )


def obra_index_from_dashboard(dashboard: dict[str, Any]) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for obra in (dashboard.get("dimensoes", {}) or {}).get("obras", []) or []:
        for centro_id in obra.get("centroCustoIds", []) or []:
            index[str(centro_id)] = obra
    return index


def faturamento_por_obra_from_dashboard(
    dashboard: dict[str, Any],
    empresa_id: str,
    inicio: str,
    fim: str,
) -> dict[str, float]:
    obras_by_centro = obra_index_from_dashboard(dashboard)
    totals: dict[str, float] = {}

    movimentos = ((dashboard.get("fatos", {}) or {}).get("movimentos", []) or [])
    for m in movimentos:
        if str(m.get("empresaId", "")) != str(empresa_id):
            continue
        if str(m.get("tipo", "")) != "Entrada":
            continue
        if _text(m.get("tipoExtrato")) != "Recebimento":
            continue

        dt = _iso_date(m.get("data"))
        if not dt:
            continue
        if inicio and dt < inicio:
            continue
        if fim and dt > fim:
            continue

        centros = m.get("centroCustoIds") or []
        obra = None
        for cid in centros:
            obra = obras_by_centro.get(str(cid))
            if obra:
                break
        if not obra:
            continue

        obra_nome = _text(obra.get("nome")) or "(Sem obra)"
        totals[obra_nome] = totals.get(obra_nome, 0.0) + float(m.get("valor") or 0.0)

    return totals


def faturamento_por_centro_from_dashboard(
    dashboard: dict[str, Any],
    empresa_id: str,
    inicio: str,
    fim: str,
) -> dict[str, float]:
    totals: dict[str, float] = {}
    movimentos = ((dashboard.get("fatos", {}) or {}).get("movimentos", []) or [])

    for m in movimentos:
        if str(m.get("empresaId", "")) != str(empresa_id):
            continue
        if str(m.get("tipo", "")) != "Entrada":
            continue
        if _text(m.get("tipoExtrato")) != "Recebimento":
            continue

        dt = _iso_date(m.get("data"))
        if not dt:
            continue
        if inicio and dt < inicio:
            continue
        if fim and dt > fim:
            continue

        cc_ids = m.get("centroCustoIds") or []
        cc_names = m.get("centroCustoNomes") or []
        if not cc_ids:
            continue

        # usa o primeiro centro como chave (padrão mais compatível com os relatórios do Sienge)
        cid = str(cc_ids[0])
        cname = _text(cc_names[0]) if len(cc_names) > 0 else ""
        key = f"{cid} - {cname}" if cname else cid
        totals[key] = totals.get(key, 0.0) + float(m.get("valor") or 0.0)

    return totals


def faturamento_por_titulo_from_dashboard(
    dashboard: dict[str, Any],
    empresa_id: str,
    inicio: str,
    fim: str,
) -> dict[str, float]:
    totals: dict[str, float] = {}
    movimentos = ((dashboard.get("fatos", {}) or {}).get("movimentos", []) or [])
    for m in movimentos:
        if str(m.get("empresaId", "")) != str(empresa_id):
            continue
        if str(m.get("tipo", "")) != "Entrada":
            continue
        if _text(m.get("tipoExtrato")) != "Recebimento":
            continue

        dt = _iso_date(m.get("data"))
        if not dt:
            continue
        if inicio and dt < inicio:
            continue
        if fim and dt > fim:
            continue

        bill_id = _text(m.get("billId"))
        if not bill_id:
            continue

        totals[bill_id] = totals.get(bill_id, 0.0) + float(m.get("valor") or 0.0)
    return totals


def faturamento_emitido_por_titulo_from_dashboard(
    dashboard: dict[str, Any],
    empresa_id: str,
    inicio: str,
    fim: str,
) -> dict[str, float]:
    totals: dict[str, float] = {}
    titulos = ((dashboard.get("fatos", {}) or {}).get("titulosReceber", []) or [])
    for t in titulos:
        if str(t.get("empresaId", "")) != str(empresa_id):
            continue
        dt = _iso_date(t.get("data"))
        if not dt:
            continue
        if inicio and dt < inicio:
            continue
        if fim and dt > fim:
            continue
        tid = _text(t.get("id"))
        if not tid:
            continue
        totals[tid] = totals.get(tid, 0.0) + float(t.get("valor") or 0.0)
    return totals


def brl(n: float) -> str:
    return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def main() -> int:
    ap = argparse.ArgumentParser(description="Compara faturamento (Sienge XLSX) vs dashboard_financeiro.json")
    ap.add_argument("--xlsx", required=True, help="Caminho do XLSX exportado do Sienge")
    ap.add_argument("--sheet", default=None, help="Nome da aba (opcional)")
    ap.add_argument("--dashboard", default=os.path.join("web", "data", "dashboard_financeiro.json"))
    ap.add_argument("--empresa", default="1")
    ap.add_argument("--inicio", default=None, help="Início ISO (YYYY-MM-DD). Se omitido, tenta usar o período do XLSX")
    ap.add_argument("--fim", default=None, help="Fim ISO (YYYY-MM-DD). Se omitido, tenta usar o período do XLSX")
    ap.add_argument(
        "--period",
        default="auto",
        choices=["auto", "args", "xlsx"],
        help="Como escolher o período: auto(usa XLSX se existir), args(obriga --inicio/--fim), xlsx(obriga período no XLSX)",
    )
    ap.add_argument("--top", type=int, default=30)
    args = ap.parse_args()

    xlsx_path = args.xlsx
    if not os.path.exists(xlsx_path):
        raise SystemExit(f"XLSX não encontrado: {xlsx_path}")

    with open(args.dashboard, "r", encoding="utf-8") as f:
        dashboard = json.load(f)

    xlsx_period = _detect_period_from_xlsx(xlsx_path, args.sheet)

    inicio = args.inicio
    fim = args.fim

    if args.period == "args":
        if not inicio or not fim:
            raise SystemExit("--period args requer --inicio e --fim")
    elif args.period == "xlsx":
        if not xlsx_period:
            raise SystemExit("Não consegui detectar o período no XLSX (campo 'Período').")
        inicio, fim = xlsx_period
    else:
        # auto
        if xlsx_period and (not inicio or not fim):
            inicio, fim = xlsx_period
        elif xlsx_period and inicio and fim and (inicio, fim) != xlsx_period:
            print(
                "(Aviso) Período informado por args difere do período do XLSX: "
                f"args={inicio}..{fim} vs xlsx={xlsx_period[0]}..{xlsx_period[1]}. "
                "Use --period xlsx para alinhar automaticamente."
            )

    # fallback: mantém compatibilidade quando nada foi passado e XLSX não tem período
    if not inicio:
        inicio = "2026-01-01"
    if not fim:
        fim = "2026-05-12"

    sheet_used, layout, rows = extract_rows_from_xlsx(xlsx_path, args.sheet, inicio, fim)

    print("=== Comparação de Faturamento ===")
    print(f"Aba XLSX usada: {sheet_used} | Layout: {layout}")
    if xlsx_period:
        print(f"Período do XLSX detectado: {xlsx_period[0]} a {xlsx_period[1]}")
    else:
        print("Período do XLSX detectado: (não encontrado)")
    print(f"Empresa: {args.empresa} | Período usado na comparação: {inicio} a {fim}")
    print("")

    if layout == "faturamento":
        # XLSX por Centro de custo / Valor bruto
        report_cc: dict[str, float] = {}
        report_titulo: dict[str, float] = {}
        report_total = 0.0
        for r in rows:
            report_total += float(r.total)
            report_cc[_norm_key(r.key)] = report_cc.get(_norm_key(r.key), 0.0) + float(r.total)
            if r.titulo:
                report_titulo[_norm_key(r.titulo)] = report_titulo.get(_norm_key(r.titulo), 0.0) + float(r.total)

        dash_emit_raw = faturamento_emitido_por_titulo_from_dashboard(dashboard, args.empresa, inicio, fim)
        dash_emit = {_norm_key(k): v for k, v in dash_emit_raw.items()}
        dash_emit_total = sum(dash_emit.values())

        print(f"Total Sienge (XLSX, Valor bruto): {brl(report_total)}")
        print(f"Total JSON (titulosReceber, emitido): {brl(dash_emit_total)}")
        print(f"Diferença total (JSON - XLSX): {brl(dash_emit_total - report_total)}")
        print("")

        def show_diffs(title: str, a_map: dict[str, float], b_map: dict[str, float], top: int) -> None:
            all_keys = sorted(set(a_map) | set(b_map))
            diffs: list[tuple[float, str, float, float]] = []
            for k in all_keys:
                a = a_map.get(k, 0.0)
                b = b_map.get(k, 0.0)
                diffs.append((abs(a - b), k, a, b))
            diffs.sort(reverse=True)

            print(title)
            shown = 0
            for d, k, a, b in diffs:
                if d < 0.005:
                    continue
                print(f"- {k} | Sienge {brl(a)} | JSON {brl(b)} | Dif {brl(b - a)}")
                shown += 1
                if shown >= top:
                    break
            if shown == 0:
                print("OK: Sem diferenças relevantes.")
            print("")

        if report_titulo:
            show_diffs(
                f"Top {args.top} diferenças por título (emitido):",
                report_titulo,
                dash_emit,
                args.top,
            )
        else:
            print("(Obs) XLSX não tem coluna 'Título'; pulando comparação por título.\n")

        # comparação secundária: recebido no caixa (movimentos) por título
        dash_receb_raw = faturamento_por_titulo_from_dashboard(dashboard, args.empresa, inicio, fim)
        dash_receb = {_norm_key(k): v for k, v in dash_receb_raw.items()}
        if report_titulo:
            show_diffs(
                f"Top {min(args.top, 15)} diferenças por título (recebido no caixa - movimentos):",
                report_titulo,
                dash_receb,
                min(args.top, 15),
            )

        # comparação secundária: por centro de custo (XLSX emitido vs recebido no caixa)
        dash_cc_raw = faturamento_por_centro_from_dashboard(dashboard, args.empresa, inicio, fim)
        dash_cc = {_norm_key(k): v for k, v in dash_cc_raw.items()}
        show_diffs(
            f"Top {min(args.top, 15)} diferenças por centro de custo (XLSX emitido vs JSON recebido):",
            report_cc,
            dash_cc,
            min(args.top, 15),
        )
    else:
        # layout antigo por grupo
        report_map: dict[str, float] = {}
        for r in rows:
            k = _norm_key(r.key)
            report_map[k] = report_map.get(k, 0.0) + float(r.total)

        dash_map_raw = faturamento_por_obra_from_dashboard(dashboard, args.empresa, inicio, fim)
        dash_map: dict[str, float] = {_norm_key(k): v for k, v in dash_map_raw.items()}

        all_keys = sorted(set(report_map) | set(dash_map))
        diffs = []
        for k in all_keys:
            a = report_map.get(k, 0.0)
            b = dash_map.get(k, 0.0)
            diffs.append((abs(a - b), k, a, b))

        diffs.sort(reverse=True)

        total_report = sum(report_map.values())
        total_dash = sum(dash_map.values())

        print(f"Total Sienge (XLSX): {brl(total_report)}")
        print(f"Total Dashboard (JSON): {brl(total_dash)}")
        print(f"Diferença total: {brl(total_dash - total_report)}")
        print("")

        print(f"Top {args.top} diferenças por grupo (normalizado):")
        shown = 0
        for d, k, a, b in diffs:
            if d < 0.005:
                continue
            print(f"- {k} | Sienge {brl(a)} | Dashboard {brl(b)} | Dif {brl(b - a)}")
            shown += 1
            if shown >= args.top:
                break

        if shown == 0:
            print("OK: Sem diferenças relevantes por grupo.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
